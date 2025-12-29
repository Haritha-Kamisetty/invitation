"""
Export Service for generating Excel, CSV, and PDF exports
Also handles CSV import and QR code generation
"""
import pandas as pd
import qrcode
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def export_guests_to_excel(event, guests):
    """
    Export guest list to Excel file
    
    Args:
        event: Event object
        guests: List of Guest objects
    
    Returns:
        BytesIO: Excel file in memory
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Guest List"
    
    # Add header
    headers = ['Name', 'Email', 'Phone', 'RSVP Status', 'Plus Ones', 'Dietary Restrictions', 'Notes', 'RSVP Time']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for guest in guests:
        ws.append([
            guest.name,
            guest.email or '',
            guest.phone or '',
            guest.rsvp_status,
            guest.plus_one_count,
            getattr(guest, 'dietary_restrictions', '') or '',
            guest.notes or '',
            guest.rsvp_time.strftime('%Y-%m-%d %H:%M') if guest.rsvp_time else ''
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def export_guests_to_csv(guests):
    """
    Export guest list to CSV format
    
    Args:
        guests: List of Guest objects
    
    Returns:
        str: CSV content
    """
    data = []
    for guest in guests:
        data.append({
            'Name': guest.name,
            'Email': guest.email or '',
            'Phone': guest.phone or '',
            'RSVP Status': guest.rsvp_status,
            'Plus Ones': guest.plus_one_count,
            'Dietary Restrictions': getattr(guest, 'dietary_restrictions', '') or '',
            'Notes': guest.notes or '',
            'RSVP Time': guest.rsvp_time.strftime('%Y-%m-%d %H:%M') if guest.rsvp_time else ''
        })
    
    df = pd.DataFrame(data)
    return df.to_csv(index=False)

def import_guests_from_csv(csv_file):
    """
    Import guests from CSV file
    
    Args:
        csv_file: File object or path to CSV
    
    Returns:
        list: List of guest dictionaries
    """
    try:
        df = pd.read_csv(csv_file)
        
        # Expected columns
        required_columns = ['Name']
        optional_columns = ['Email', 'Phone', 'Dietary Restrictions', 'Notes']
        
        # Validate required columns
        if 'Name' not in df.columns:
            raise ValueError("CSV must contain 'Name' column")
        
        guests = []
        for _, row in df.iterrows():
            guest_data = {
                'name': str(row['Name']).strip(),
                'email': str(row.get('Email', '')).strip() if pd.notna(row.get('Email')) else None,
                'phone': str(row.get('Phone', '')).strip() if pd.notna(row.get('Phone')) else None,
                'dietary_restrictions': str(row.get('Dietary Restrictions', '')).strip() if pd.notna(row.get('Dietary Restrictions')) else None,
                'notes': str(row.get('Notes', '')).strip() if pd.notna(row.get('Notes')) else None
            }
            
            # Skip empty rows
            if guest_data['name']:
                guests.append(guest_data)
        
        return guests
        
    except Exception as e:
        raise ValueError(f"Error importing CSV: {str(e)}")

def generate_qr_code(url, event_id, qr_folder):
    """
    Generate QR code for event invitation URL
    
    Args:
        url: URL to encode in QR code
        event_id: Event ID for filename
        qr_folder: Folder to save QR code
    
    Returns:
        str: Path to generated QR code image
    """
    try:
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Ensure folder exists
        os.makedirs(qr_folder, exist_ok=True)
        
        # Save image
        filename = f"qr_{event_id}.png"
        filepath = os.path.join(qr_folder, filename)
        img.save(filepath)
        
        return filepath
        
    except Exception as e:
        print(f"Error generating QR code: {str(e)}")
        return None

def get_event_statistics(event, guests):
    """
    Generate comprehensive statistics for an event
    
    Args:
        event: Event object
        guests: List of Guest objects
    
    Returns:
        dict: Statistics dictionary
    """
    total_guests = len(guests)
    rsvp_yes = sum(1 for g in guests if g.rsvp_status == 'Yes')
    rsvp_no = sum(1 for g in guests if g.rsvp_status == 'No')
    rsvp_maybe = sum(1 for g in guests if g.rsvp_status == 'Maybe')
    rsvp_pending = sum(1 for g in guests if g.rsvp_status == 'Pending')
    
    total_attendees = rsvp_yes + sum(g.plus_one_count for g in guests if g.rsvp_status == 'Yes')
    
    # Dietary restrictions summary
    dietary_restrictions = {}
    for guest in guests:
        if hasattr(guest, 'dietary_restrictions') and guest.dietary_restrictions:
            restrictions = guest.dietary_restrictions.split(',')
            for restriction in restrictions:
                restriction = restriction.strip()
                if restriction:
                    dietary_restrictions[restriction] = dietary_restrictions.get(restriction, 0) + 1
    
    return {
        'total_guests': total_guests,
        'rsvp_yes': rsvp_yes,
        'rsvp_no': rsvp_no,
        'rsvp_maybe': rsvp_maybe,
        'rsvp_pending': rsvp_pending,
        'total_attendees': total_attendees,
        'response_rate': round((total_guests - rsvp_pending) / total_guests * 100, 1) if total_guests > 0 else 0,
        'dietary_restrictions': dietary_restrictions
    }
